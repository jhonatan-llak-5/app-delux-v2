"""
Bulk import de productos desde xlsx.

Flujo:
  1. build_template_xlsx() -> bytes  (descarga inicial)
  2. parse_xlsx(file) -> rows[]      (parser tolerante)
  3. validate_rows(rows) -> rows[]   (status: ok | warning | error)
  4. extract_zip_images(zip_file)    (opcional, auto-match por slug)
  5. commit_rows(rows, image_map)    (crea Product + variants + stock + images)
"""
from __future__ import annotations

import io
import os
import re
import zipfile
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from django.conf import settings
from django.db import transaction
from django.utils.text import slugify

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from apps.brands.models import Brand
from apps.branches.models import Branch
from apps.categories.models import Category
from apps.tenants.models import Tenant

from .models import Product, ProductImage, ProductStatus, ProductTag, ProductKind


# ============================================================
# Definición de columnas
# ============================================================
COLUMNS: List[Dict[str, Any]] = [
    {'key': 'name',                'label': 'Nombre',              'required': True,  'width': 32},
    {'key': 'brand_slug',          'label': 'Marca',               'required': True,  'width': 18},
    {'key': 'category_slug',       'label': 'Categoría',           'required': True,  'width': 20},
    {'key': 'kind',                'label': 'Tipo',                'required': False, 'width': 12},
    {'key': 'base_price',          'label': 'Precio base',         'required': True,  'width': 12},
    {'key': 'compare_at_price',    'label': 'Precio comparado',    'required': False, 'width': 14},
    {'key': 'gender',              'label': 'Género',              'required': False, 'width': 12},
    {'key': 'status',              'label': 'Estado',              'required': False, 'width': 14},
    {'key': 'tag',                 'label': 'Tag',                 'required': False, 'width': 12},
    {'key': 'is_featured',         'label': 'Destacado',           'required': False, 'width': 10},
    {'key': 'short_description',   'label': 'Descripción corta',   'required': False, 'width': 40},
    {'key': 'description',         'label': 'Descripción larga',   'required': False, 'width': 60},
    {'key': 'sizes',               'label': 'Tallas (sep |)',      'required': False, 'width': 22},
    {'key': 'colors',              'label': 'Colores (sep |)',     'required': False, 'width': 22},
    {'key': 'stock_per_variant',   'label': 'Stock por variante',  'required': False, 'width': 12},
    {'key': 'cost',                'label': 'Costo (compra)',      'required': False, 'width': 12},
    {'key': 'supplier',            'label': 'Proveedor',           'required': False, 'width': 22},
    {'key': 'main_image_url',      'label': 'URL imagen principal','required': False, 'width': 40},
    {'key': 'extra_images_urls',   'label': 'URLs extras (sep |)', 'required': False, 'width': 40},
]

GENDER_CHOICES  = ['UNISEX', 'MEN', 'WOMEN', 'KIDS']
STATUS_CHOICES  = [s.value for s in ProductStatus]
TAG_CHOICES     = [''] + [t.value for t in ProductTag]
BOOL_CHOICES    = ['true', 'false']
KIND_CHOICES    = [k.value for k in ProductKind]

HEADER_FILL = PatternFill(start_color='FF1E40AF', end_color='FF1E40AF', fill_type='solid')
HEADER_FONT = Font(color='FFFFFFFF', bold=True, size=11)
SUBHEAD_FILL = PatternFill(start_color='FFF1F5F9', end_color='FFF1F5F9', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='FFE2E8F0'),
    right=Side(style='thin', color='FFE2E8F0'),
    top=Side(style='thin', color='FFE2E8F0'),
    bottom=Side(style='thin', color='FFE2E8F0'),
)


# ============================================================
# 1. Plantilla xlsx con dropdowns
# ============================================================
def build_template_xlsx(tenant: Optional[Tenant] = None) -> bytes:
    """Genera una plantilla .xlsx con headers + dropdowns para género/estado/tag/marca/categoría/sucursal."""
    wb = Workbook()

    # ---- Hoja 1: Productos ----
    ws = wb.active
    ws.title = 'Productos'

    # Headers
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col['label'])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col['width']
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    # Ejemplo en row 2
    example = {
        'name': 'Nike Air Force 1 \'07',
        'brand_slug': 'Nike',
        'category_slug': 'Zapatillas',
        'kind': 'CALZADO',
        'base_price': 119.90,
        'compare_at_price': 139.90,
        'gender': 'UNISEX',
        'status': 'PUBLISHED',
        'tag': 'NEW',
        'is_featured': 'true',
        'short_description': 'Clásico icónico, cuero premium.',
        'description': 'Descripción larga del producto…',
        'sizes': '38|39|40|41|42|43',
        'colors': 'Blanco|Negro',
        'stock_per_variant': 10,
        'cost': 45.00,
        'supplier': 'Proveedor Ejemplo',
        'main_image_url': 'https://ejemplo.com/af1-main.jpg',
        'extra_images_urls': 'https://ejemplo.com/af1-2.jpg|https://ejemplo.com/af1-3.jpg',
    }
    for col_idx, col in enumerate(COLUMNS, start=1):
        c = ws.cell(row=2, column=col_idx, value=example.get(col['key'], ''))
        c.fill = SUBHEAD_FILL
        c.font = Font(italic=True, color='FF64748B')
        c.border = THIN_BORDER

    # Validaciones (data validation drop-downs)
    last_row = 1000
    # Cargar slugs reales si hay tenant
    brand_names = list(Brand.objects.filter(is_active=True).values_list('name', flat=True)) if Brand.objects.exists() else []
    category_names = list(Category.objects.filter(is_active=True).values_list('name', flat=True)) if Category.objects.exists() else []
    branch_codes = list(Branch.objects.filter(is_active=True).values_list('code', flat=True)) if Branch.objects.exists() else []

    def col_letter(key: str) -> str:
        return get_column_letter(_col_index(key))

    def add_dv(formula: str, key: str):
        dv = DataValidation(type='list', formula1=formula, allow_blank=True)
        dv.error = 'Valor inválido. Selecciona uno de la lista.'
        dv.errorTitle = 'Valor no permitido'
        ws.add_data_validation(dv)
        letter = col_letter(key)
        dv.add(f'{letter}3:{letter}{last_row}')

    add_dv(f'"{",".join(GENDER_CHOICES)}"', 'gender')
    add_dv(f'"{",".join(STATUS_CHOICES)}"', 'status')
    add_dv(f'"{",".join([t for t in TAG_CHOICES if t])}"', 'tag')
    add_dv(f'"{",".join(BOOL_CHOICES)}"', 'is_featured')
    add_dv(f'"{",".join(KIND_CHOICES)}"', 'kind')

    # ---- Hoja 2: Catálogos (referencia) ----
    cat_ws = wb.create_sheet('Catálogos')
    cat_ws['A1'] = 'Marcas disponibles'
    cat_ws['B1'] = 'Categorías disponibles'
    for c in ('A1', 'B1'):
        cat_ws[c].fill = HEADER_FILL
        cat_ws[c].font = HEADER_FONT
    for i, v in enumerate(brand_names, start=2):    cat_ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(category_names, start=2): cat_ws.cell(row=i, column=2, value=v)
    for col in ('A', 'B', 'C'):
        cat_ws.column_dimensions[col].width = 24

    # ---- Hoja 3: Instrucciones ----
    help_ws = wb.create_sheet('Instrucciones')
    help_ws['A1'] = 'Importador masivo de productos — Delux'
    help_ws['A1'].font = Font(bold=True, size=14, color='FF1E40AF')
    instructions = [
        '',
        '1. Llena la hoja "Productos" con una fila por producto.',
        '2. Columnas obligatorias: Nombre, Marca, Categoría, Precio base.',
        '3. En Marca y Categoría escribe el NOMBRE tal como aparece en la hoja "Catálogos".',
        '4. Para variantes: separa tallas y colores con barra vertical |  (ej: 38|39|40).',
        '   Se crean todas las combinaciones (talla × color).',
        '5. Stock: indica "Stock por variante". La SUCURSAL destino se elige al subir el',
        '   archivo (paso 2), no en el Excel. El stock se cargará en la(s) sucursal(es)',
        '   que selecciones. Costo = costo de compra; Proveedor = nombre (se crea solo).',
        '   - Tipo: CALZADO, ROPA, GORRA, MOCHILA, BISUTERIA, ACCESORIO u OTRO.',
        '   - El código interno (P########) se genera automáticamente por variante.',
        '6. Imágenes (dos formas):',
        '   a) Por URL: llena "URL imagen principal" y "URLs extras" (separadas por |).',
        '   b) Por ZIP: sube un .zip donde cada imagen se llame igual que el NOMBRE del',
        '      producto en minúsculas y con guiones. Ej: para "Nike Air Force 1" usa',
        '      nike-air-force-1.jpg (principal) y nike-air-force-1-2.jpg, -3.jpg (extras).',
        '7. Borra la fila 2 (ejemplo) antes de subir, o se creará como producto.',
    ]
    for i, line in enumerate(instructions, start=2):
        help_ws.cell(row=i, column=1, value=line)
    help_ws.column_dimensions['A'].width = 90

    # Serializar
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _col_index(key: str) -> int:
    for i, col in enumerate(COLUMNS, start=1):
        if col['key'] == key:
            return i
    raise KeyError(f'columna desconocida: {key}')


# ============================================================
# 2. Parseo del xlsx subido
# ============================================================
def parse_xlsx(file_obj) -> List[Dict[str, Any]]:
    """Lee la hoja 'Productos' (o la primera) y devuelve filas como dicts."""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb['Productos'] if 'Productos' in wb.sheetnames else wb.worksheets[0]

    rows: List[Dict[str, Any]] = []
    iterator = ws.iter_rows(min_row=2, values_only=True)
    for idx, raw in enumerate(iterator, start=2):
        if not raw or all(c is None or str(c).strip() == '' for c in raw):
            continue
        # Mapear según COLUMNS
        row: Dict[str, Any] = {'__row__': idx}
        for i, col in enumerate(COLUMNS):
            if i < len(raw):
                v = raw[i]
                if isinstance(v, str):
                    v = v.strip()
                row[col['key']] = v if v not in (None, '') else None
            else:
                row[col['key']] = None
        rows.append(row)
    return rows


# ============================================================
# 3. Validación
# ============================================================
def validate_rows(rows: List[Dict[str, Any]], tenant: Optional[Tenant] = None) -> List[Dict[str, Any]]:
    """Anota cada fila con status (ok|warning|error) + errors[] + warnings[]."""
    # Pre-cargar catálogos (acotados al tenant), aceptando nombre o slug.
    b_qs = Brand.objects.filter(tenant=tenant) if tenant else Brand.objects.all()
    c_qs = Category.objects.filter(tenant=tenant) if tenant else Category.objects.all()
    brands: Dict[str, Any] = {}
    for b in b_qs:
        brands[b.slug.lower()] = b
        brands[(b.name or '').strip().lower()] = b
    categories: Dict[str, Any] = {}
    for c in c_qs:
        categories[c.slug.lower()] = c
        categories[(c.name or '').strip().lower()] = c

    # Slugs ya existentes en DB (del tenant)
    existing_slugs = set(
        (Product.objects.filter(tenant=tenant) if tenant else Product.objects).values_list('slug', flat=True))
    # Para detectar duplicados dentro del propio archivo
    seen_in_file: Dict[str, int] = {}

    results: List[Dict[str, Any]] = []
    for row in rows:
        errors: List[str] = []
        warnings: List[str] = []

        # Nombre
        name = row.get('name')
        if not name:
            errors.append('Nombre es obligatorio.')

        # Slug (se genera automáticamente desde el nombre; se hace único).
        base_slug = slugify(name)[:170] if name else ''
        slug = base_slug
        n = 2
        while slug and (slug in existing_slugs or slug in seen_in_file):
            slug = f'{base_slug}-{n}'
            n += 1
        row['slug'] = slug
        if slug:
            seen_in_file[slug] = row['__row__']

        # Marca (por nombre o slug)
        brand_in = (row.get('brand_slug') or '').strip()
        if not brand_in:
            errors.append('Marca es obligatoria.')
        elif brand_in.lower() not in brands:
            errors.append(f'Marca "{brand_in}" no existe. Ver hoja Catálogos.')
        else:
            row['brand_slug'] = brands[brand_in.lower()].slug

        # Categoría (por nombre o slug)
        cat_in = (row.get('category_slug') or '').strip()
        if not cat_in:
            errors.append('Categoría es obligatoria.')
        elif cat_in.lower() not in categories:
            errors.append(f'Categoría "{cat_in}" no existe. Ver hoja Catálogos.')
        else:
            row['category_slug'] = categories[cat_in.lower()].slug

        # Precio
        price = row.get('base_price')
        try:
            row['base_price'] = Decimal(str(price)) if price is not None else None
            if row['base_price'] is None or row['base_price'] <= 0:
                errors.append('Precio base debe ser > 0.')
        except (InvalidOperation, TypeError):
            errors.append(f'Precio base inválido: "{price}".')

        # Precio comparado
        cap = row.get('compare_at_price')
        if cap is not None:
            try:
                row['compare_at_price'] = Decimal(str(cap))
            except (InvalidOperation, TypeError):
                warnings.append(f'Precio comparado inválido "{cap}" — se ignora.')
                row['compare_at_price'] = None

        # Género
        gender = (row.get('gender') or 'UNISEX').upper()
        if gender not in GENDER_CHOICES:
            warnings.append(f'Género "{gender}" inválido — se usa UNISEX.')
            gender = 'UNISEX'
        row['gender'] = gender

        # Status
        status = (row.get('status') or 'DRAFT').upper()
        if status not in STATUS_CHOICES:
            warnings.append(f'Estado "{status}" inválido — se usa DRAFT.')
            status = 'DRAFT'
        row['status'] = status

        # Tag
        tag = (row.get('tag') or '').upper()
        if tag and tag not in [t for t in TAG_CHOICES if t]:
            warnings.append(f'Tag "{tag}" inválido — se ignora.')
            tag = ''
        row['tag'] = tag

        # is_featured
        feat = row.get('is_featured')
        row['is_featured'] = str(feat).strip().lower() in ('true', '1', 'sí', 'si', 'yes')

        # Tipo de producto
        kind = (row.get('kind') or 'OTRO').upper()
        if kind not in KIND_CHOICES:
            warnings.append(f'Tipo "{kind}" inválido — se usa OTRO.')
            kind = 'OTRO'
        row['kind'] = kind

        # Costo
        cost = row.get('cost')
        if cost is not None:
            try:
                row['cost'] = Decimal(str(cost))
            except (InvalidOperation, TypeError):
                warnings.append(f'Costo "{cost}" inválido — se usa 0.')
                row['cost'] = Decimal('0')
        else:
            row['cost'] = Decimal('0')

        # Proveedor (texto libre, opcional)
        row['supplier'] = (str(row.get('supplier')).strip() if row.get('supplier') else '')

        # Variantes: parsear listas
        sizes = _parse_pipe_list(row.get('sizes'))
        colors = _parse_pipe_list(row.get('colors'))
        row['sizes'] = sizes
        row['colors'] = colors
        if not sizes and not colors:
            warnings.append('Sin tallas ni colores — se creará variante única.')

        # Stock
        stock = row.get('stock_per_variant')
        if stock is not None:
            try:
                row['stock_per_variant'] = int(stock)
            except (TypeError, ValueError):
                warnings.append(f'Stock "{stock}" inválido — se ignora.')
                row['stock_per_variant'] = 0
        else:
            row['stock_per_variant'] = 0

        # La sucursal destino del stock se elige al subir (no en el Excel).

        # URLs imagen
        row['extra_images_urls'] = _parse_pipe_list(row.get('extra_images_urls'))

        # Determinar status final
        if errors:
            row['_status'] = 'error'
        elif warnings:
            row['_status'] = 'warning'
        else:
            row['_status'] = 'ok'
        row['_errors'] = errors
        row['_warnings'] = warnings
        results.append(row)

    return results


def _parse_pipe_list(value: Any) -> List[str]:
    if not value:
        return []
    return [p.strip() for p in str(value).split('|') if p.strip()]


# ============================================================
# 4. Imágenes desde ZIP (auto-match por slug)
# ============================================================
def extract_zip_images(zip_file_obj, dest_subdir: str = 'products/bulk') -> Dict[str, List[str]]:
    """
    Extrae imágenes de un ZIP y las guarda en MEDIA_ROOT/{dest_subdir}/{timestamp}/.

    Retorna mapping slug_prefix -> [url1, url2, ...] (ordenadas).
    Asume nombres como  nike-air-force-1.jpg  o  nike-air-force-1-2.jpg.
    """
    image_map: Dict[str, List[Tuple[int, str]]] = {}

    if not zip_file_obj:
        return {}

    ts = datetime.utcnow().strftime('%Y%m%d-%H%M%S')
    target_dir = os.path.join(settings.MEDIA_ROOT, dest_subdir, ts)
    os.makedirs(target_dir, exist_ok=True)

    media_url_base = settings.MEDIA_URL.rstrip('/') + '/' + dest_subdir + '/' + ts + '/'

    with zipfile.ZipFile(zip_file_obj) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            filename = os.path.basename(info.filename)
            if not filename or filename.startswith('.'):
                continue
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ('.jpg', '.jpeg', '.png', '.webp', '.gif', '.avif'):
                continue

            # Sanitizar nombre para evitar path traversal
            safe_name = re.sub(r'[^A-Za-z0-9._-]', '_', filename)
            target_path = os.path.join(target_dir, safe_name)
            with zf.open(info) as src, open(target_path, 'wb') as dst:
                dst.write(src.read())

            url = media_url_base + safe_name

            # Extraer slug + order: "nike-air-force-1-2.jpg" -> slug="nike-air-force-1", order=2
            stem = os.path.splitext(safe_name)[0]
            m = re.match(r'^(.*?)(?:-(\d+))?$', stem)
            slug_prefix = (m.group(1) if m else stem).lower()
            order = int(m.group(2)) if (m and m.group(2)) else 0
            image_map.setdefault(slug_prefix, []).append((order, url))

    # Ordenar y devolver solo URLs
    return {slug: [u for _, u in sorted(lst, key=lambda x: x[0])] for slug, lst in image_map.items()}


# ============================================================
# 5. Commit — crear productos
# ============================================================
@transaction.atomic
def commit_rows(
    validated_rows: List[Dict[str, Any]],
    image_map: Optional[Dict[str, List[str]]] = None,
    tenant: Optional[Tenant] = None,
    branch_codes: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Crea productos válidos; ignora filas en error. Retorna resumen."""
    from django.utils import timezone
    from apps.variants.models import Variant
    from apps.inventory.models import Stock, StockMovement, Supplier, Reception, ReceptionItem
    from apps.inventory.services import next_sku_number

    if tenant is None:
        tenant = Tenant.objects.filter(is_active=True).first()

    seq = next_sku_number(tenant)
    receptions: Dict[int, Any] = {}

    def _get_reception(branch, supplier):
        rec = receptions.get(branch.id)
        if rec is None:
            rec = Reception.objects.create(
                tenant=tenant, branch=branch, supplier=supplier,
                status='COMMITTED', committed_at=timezone.now(),
                note='Importación masiva',
            )
            rec.code = f'REC-{timezone.now():%Y%m%d}-{rec.pk:04d}'
            rec.save(update_fields=['code'])
            receptions[branch.id] = rec
        elif supplier and rec.supplier_id is None:
            rec.supplier = supplier
            rec.save(update_fields=['supplier'])
        return rec

    image_map = image_map or {}
    brands = {b.slug: b for b in (Brand.objects.filter(tenant=tenant) if tenant else Brand.objects.all())}
    categories = {c.slug: c for c in (Category.objects.filter(tenant=tenant) if tenant else Category.objects.all())}
    branch_qs = Branch.objects.filter(tenant=tenant) if tenant else Branch.objects.all()
    branches = {b.code: b for b in branch_qs}
    # Sucursales destino elegidas al subir (el stock se replica a cada una).
    dest_branches = [branches[c] for c in (branch_codes or []) if c in branches]
    existing_slugs = set(
        (Product.objects.filter(tenant=tenant) if tenant else Product.objects).values_list('slug', flat=True))

    created: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []

    for row in validated_rows:
        if row.get('_status') == 'error':
            skipped.append({'row': row['__row__'], 'reason': '; '.join(row.get('_errors', []))})
            continue
        # slug ya viene único desde validate; si por carrera existiera, se ajusta.
        while row['slug'] in existing_slugs:
            row['slug'] = row['slug'] + '-x'

        product = Product.objects.create(
            tenant=tenant,
            name=row['name'],
            slug=row['slug'],
            short_description=row.get('short_description') or '',
            description=row.get('description') or '',
            brand=brands[row['brand_slug']],
            category=categories[row['category_slug']],
            base_price=row['base_price'],
            compare_at_price=row.get('compare_at_price'),
            gender=row['gender'],
            status=row['status'],
            tag=row['tag'],
            is_featured=row['is_featured'],
            main_image_url=row.get('main_image_url') or '',
            kind=row.get('kind') or 'OTRO',
        )
        existing_slugs.add(product.slug)

        # Imágenes: URLs explícitas + ZIP matcheado
        urls = list(row.get('extra_images_urls', []))
        zip_urls = image_map.get(product.slug, [])
        # Si no hay main_image_url y el ZIP tiene la primera, usarla
        if not product.main_image_url and zip_urls:
            product.main_image_url = zip_urls[0]
            product.save(update_fields=['main_image_url'])
            zip_urls = zip_urls[1:]
        urls.extend(zip_urls)

        for idx, url in enumerate(urls):
            ProductImage.objects.create(product=product, url=url, sort_order=idx + 1, is_main=False)

        # Variantes (combinaciones talla × color)
        sizes = row.get('sizes') or ['']
        colors = row.get('colors') or ['']
        stock_qty = row.get('stock_per_variant') or 0

        cost = row.get('cost') or Decimal('0')
        sup_name = (row.get('supplier') or '').strip()
        supplier_obj = (Supplier.objects.get_or_create(tenant=tenant, name=sup_name)[0]
                        if sup_name else None)

        for s in sizes:
            for c in colors:
                sku = f'P{seq:08d}'
                seq += 1
                variant = Variant.objects.create(
                    tenant=tenant, product=product,
                    sku=sku, size=s, color=c, cost=cost,
                )
                # Stock inicial: se carga en CADA sucursal destino seleccionada.
                if stock_qty > 0:
                    for branch in dest_branches:
                        stock = Stock.objects.create(
                            tenant=tenant, variant=variant, branch=branch,
                            quantity=stock_qty,
                        )
                        StockMovement.objects.create(
                            tenant=tenant, stock=stock, type=StockMovement.TYPE_IN,
                            quantity=stock_qty, note='Importación masiva',
                        )
                        rec = _get_reception(branch, supplier_obj)
                        ReceptionItem.objects.create(
                            tenant=tenant, reception=rec, variant=variant,
                            quantity=stock_qty, unit_cost=cost,
                        )

        created.append({'row': row['__row__'], 'id': product.id, 'slug': product.slug, 'name': product.name})

    return {
        'created_count': len(created),
        'skipped_count': len(skipped),
        'created': created,
        'skipped': skipped,
    }
